#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from read_excel_MIK import read_excel_MIK
from read_excel_Seichitech import read_excel_Seichitech
from openpyxl import load_workbook
import pandas

class read_write_excel:

    """
    For read/write and parse excel file
    """

    def __init__(self, filename):

        '''
        support both MIK and Seichitech platform
        '''

        self.filename = filename
        fd = pandas.ExcelFile(self.filename)
        sheet_names = fd.sheet_names
        fd.close()
        if "Result Data" in sheet_names and "Graph" in sheet_names and "Raw Data" in sheet_names:
            self.fd = read_excel_MIK(self.filename)
        else:
            self.fd = read_excel_Seichitech(self.filename)

    def read_config(self):

        '''
        read boundary, max_x, max_y information
        '''

        return self.fd.read_config()

    def read_report_time(self):

        '''
        read report time for line test
        '''

        return self.fd.read_report_time()

    def read_target(self):

        '''
        get target from excel
        '''

        return self.fd.read_target()

    def read_measure(self):

        '''
        get measure data from excel
        '''

        return self.fd.read_measure()

    def read_target_and_measure(self):

        '''
        read target and measure at the same time
        '''

        target_list = self.read_target()
        measure_list_mm, measure_list_pixel = self.read_measure()

        return target_list, measure_list_mm, measure_list_pixel

    def write_excel(self, write_data):

        '''
        write output to new sheet
        '''

        df = pandas.DataFrame(write_data)
        book = load_workbook(self.filename)
        sheet_names = book.sheetnames
        for name in sheet_names:
            if "analysis_output" in name:
                book.remove(book[name])
        writer = pandas.ExcelWriter(self.filename, engine = 'openpyxl')
        writer.book = book
        df.to_excel(writer, sheet_name='analysis_output', index=False)
        work_sheet = book["analysis_output"]
        for col in work_sheet.columns:
            max_length = 0
            column = col[0].column # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            work_sheet.column_dimensions[column].width = adjusted_width
        writer.save()
        writer.close()

    def destroy(self):

        '''
        destroy excel fd
        '''

        self.fd.destroy()

if __name__ == "__main__":

    """
    This is for test purpose
    """

    fd = read_write_excel("../H_Line/Result.xlsx")
    test_type, max_x, max_y, boundary_range = fd.read_config()
    target_list, measure_list = fd.read_target_and_measure()
    # print("This is %s, max_x = %f, max_y = %f, boundary_range = %f" % ( test_type, max_x, max_y, boundary_range ))
    # for i in range(len(target_list)):
    #     print("\nNO.%d target line ---------------------- (%f, %f) -> (%f, %f)" % ( i + 1, target_list[i][0], target_list[i][1], target_list[i][2], target_list[i][3] ))
    #     for j in range(len(measure_list[i])):
    #         print("\tNO.%d measured point ---------------------------- (%f, %f)" % ( j + 1, measure_list[i][j][0], measure_list[i][j][1] ))
    #     print("\n")
    test_dict = {'a':[1,2,3], 'b':[2,3,4], 'c':[3,4,5]}
    fd.write_excel(test_dict)

    fd = read_write_excel("../POINTS/20180918175024.xlsx")
    test_type, max_x, max_y, boundary_range = fd.read_config()
    target_list, measure_list = fd.read_target_and_measure()
    # print("This is %s, max_x = %f, max_y = %f, boundary_range = %f" % ( test_type, max_x, max_y, boundary_range ))
    # for i in range(len(target_list)):
    #     print("\nNO.%d target point ---------------------- (%f, %f)" % ( i + 1, target_list[i][0], target_list[i][1] ))
    #     for j in range(len(measure_list[i])):
    #         print("\tRepeat %d" % ( j + 1 ))
    #         for k in range(len(measure_list[i][j])):
    #             print("\t\tNO.%d measured point ---------------------------- (%f, %f)" % ( k + 1, measure_list[i][j][k][0], measure_list[i][j][k][1] ))
    #         print("\n")
